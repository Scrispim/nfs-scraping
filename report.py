import io
import re
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Font,
    PatternFill,
    Border,
    Side,
    numbers,
)
from openpyxl.utils import get_column_letter


# Paleta de cores
_GREEN_DARK = "1A5276"
_GREEN_MID = "1ABC9C"
_GREEN_LIGHT = "D5F5E3"
_HEADER_FONT_COLOR = "FFFFFF"
_ZEBRA = "F2F9F6"


def _thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _parse_value(val: str) -> float:
    """Converte 'R$ 1.234,56' ou '1.234,56' para float."""
    cleaned = re.sub(r"[^\d,]", "", str(val)).replace(",", ".")
    try:
        return float(cleaned)
    except ValueError:
        return 0.0


def _to_dataframe(records: list[dict]) -> pd.DataFrame:
    df = pd.DataFrame(records)
    if df.empty:
        return df

    df["Valor"] = df["Preço Serviço (R$)"].apply(_parse_value)
    df["Geração"] = pd.to_datetime(df["Geração"], dayfirst=True, errors="coerce")
    return df


def generate_report(records: list[dict], data_inicial: str, data_final: str) -> bytes:
    """
    Gera um arquivo Excel em memória com duas abas:
      - Resumo: totalizadores, agrupamentos por situação e por competência
      - Detalhes: todos os registros linha a linha
    Retorna os bytes do arquivo .xlsx.
    """
    df = _to_dataframe(records)
    buf = io.BytesIO()
    wb = Workbook()

    _build_summary(wb, df, data_inicial, data_final)
    _build_details(wb, df)

    wb.save(buf)
    return buf.getvalue()


# ------------------------------------------------------------------
# Aba Resumo
# ------------------------------------------------------------------

def _build_summary(wb: Workbook, df: pd.DataFrame, data_inicial: str, data_final: str):
    ws = wb.active
    ws.title = "Resumo"
    ws.sheet_view.showGridLines = False

    row = 1

    # Título principal
    ws.merge_cells(f"A{row}:F{row}")
    c = ws[f"A{row}"]
    c.value = "NFS-e — Relatório de Notas Emitidas"
    c.font = Font(bold=True, size=14, color=_HEADER_FONT_COLOR)
    c.fill = PatternFill("solid", fgColor=_GREEN_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 30
    row += 1

    # Período
    ws.merge_cells(f"A{row}:F{row}")
    c = ws[f"A{row}"]
    c.value = f"Período: {data_inicial}  →  {data_final}"
    c.font = Font(italic=True, size=10, color="555555")
    c.alignment = Alignment(horizontal="center")
    row += 2

    if df.empty:
        ws[f"A{row}"] = "Nenhuma nota encontrada no período."
        return

    total_notas = len(df)
    total_valor = df["Valor"].sum()

    # Cards de totalizadores
    _write_card(ws, row, "A", "Total de Notas", total_notas, fmt="inteiro")
    _write_card(ws, row, "D", "Valor Total (R$)", total_valor, fmt="moeda")
    row += 4

    # Agrupamento por Situação
    row = _write_group_table(
        ws, df, row,
        title="Por Situação",
        group_col="Situação",
    )
    row += 1

    # Agrupamento por Competência
    row = _write_group_table(
        ws, df, row,
        title="Por Competência",
        group_col="Competência",
    )

    # Larguras
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16

    # Rodapé
    last = ws.max_row + 2
    ws.merge_cells(f"A{last}:F{last}")
    c = ws[f"A{last}"]
    c.value = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    c.font = Font(italic=True, size=9, color="999999")
    c.alignment = Alignment(horizontal="right")


def _write_card(ws, row: int, col: str, label: str, value, fmt: str):
    col_idx = ord(col) - ord("A")
    label_col = get_column_letter(col_idx + 1)
    val_col = get_column_letter(col_idx + 2)

    # Label
    lc = ws[f"{label_col}{row}"]
    lc.value = label
    lc.font = Font(bold=True, size=10, color=_HEADER_FONT_COLOR)
    lc.fill = PatternFill("solid", fgColor=_GREEN_MID)
    lc.alignment = Alignment(horizontal="center", vertical="center")
    lc.border = _thin_border()
    ws.row_dimensions[row].height = 22

    # Value
    vc = ws[f"{label_col}{row + 1}"]
    ws.merge_cells(f"{label_col}{row + 1}:{val_col}{row + 2}")
    vc.value = value
    vc.font = Font(bold=True, size=16, color=_GREEN_DARK)
    vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.border = _thin_border()
    if fmt == "moeda":
        vc.number_format = '#,##0.00'


def _write_group_table(ws, df: pd.DataFrame, row: int, title: str, group_col: str) -> int:
    # Cabeçalho do grupo
    ws.merge_cells(f"A{row}:C{row}")
    c = ws[f"A{row}"]
    c.value = title
    c.font = Font(bold=True, size=11, color=_HEADER_FONT_COLOR)
    c.fill = PatternFill("solid", fgColor=_GREEN_DARK)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.border = _thin_border()
    ws.row_dimensions[row].height = 22
    row += 1

    # Cabeçalhos da tabela
    headers = [group_col, "Qtd. Notas", "Valor Total (R$)"]
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=i + 1, value=h)
        c.font = Font(bold=True, size=10, color=_HEADER_FONT_COLOR)
        c.fill = PatternFill("solid", fgColor=_GREEN_MID)
        c.alignment = Alignment(horizontal="center")
        c.border = _thin_border()
    row += 1

    grouped = df.groupby(group_col, dropna=False).agg(
        qtd=("Valor", "count"),
        total=("Valor", "sum"),
    ).reset_index()

    for idx, r in grouped.iterrows():
        fill = PatternFill("solid", fgColor=_ZEBRA) if idx % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for col_i, val in enumerate([r[group_col], r["qtd"], r["total"]]):
            c = ws.cell(row=row, column=col_i + 1, value=val)
            c.fill = fill
            c.border = _thin_border()
            c.alignment = Alignment(horizontal="center")
            if col_i == 2:
                c.number_format = '#,##0.00'
        row += 1

    return row


# ------------------------------------------------------------------
# Aba Detalhes
# ------------------------------------------------------------------

def _build_details(wb: Workbook, df: pd.DataFrame):
    ws = wb.create_sheet("Detalhes")
    ws.sheet_view.showGridLines = False

    headers = ["Geração", "Emitida Para", "Competência", "Município Emissor", "Preço Serviço (R$)", "Situação"]
    col_widths = [14, 40, 14, 20, 20, 16]

    # Cabeçalho
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, size=10, color=_HEADER_FONT_COLOR)
        c.fill = PatternFill("solid", fgColor=_GREEN_DARK)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _thin_border()
        ws.column_dimensions[get_column_letter(i)].width = col_widths[i - 1]
    ws.row_dimensions[1].height = 28

    if df.empty:
        return

    # Dados
    display_cols = ["Geração", "Emitida Para", "Competência", "Município Emissor", "Preço Serviço (R$)", "Situação"]
    for row_idx, (_, row) in enumerate(df[display_cols].iterrows(), start=2):
        fill = PatternFill("solid", fgColor=_ZEBRA) if row_idx % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for col_idx, val in enumerate(row, start=1):
            c = ws.cell(row=row_idx, column=col_idx)
            # Data formatada
            if col_idx == 1 and pd.notna(val) and isinstance(val, pd.Timestamp):
                c.value = val.to_pydatetime()
                c.number_format = "DD/MM/YYYY"
            elif col_idx == 5:
                c.value = _parse_value(str(val))
                c.number_format = '#,##0.00'
            else:
                c.value = str(val) if pd.notna(val) else ""
            c.fill = fill
            c.border = _thin_border()
            c.alignment = Alignment(horizontal="center", vertical="center")

    # Linha de total
    total_row = ws.max_row + 1
    c = ws.cell(row=total_row, column=4, value="TOTAL")
    c.font = Font(bold=True)
    c.fill = PatternFill("solid", fgColor=_GREEN_LIGHT)
    c.border = _thin_border()
    c.alignment = Alignment(horizontal="right")

    total_c = ws.cell(row=total_row, column=5)
    total_c.value = df["Valor"].sum()
    total_c.font = Font(bold=True)
    total_c.number_format = '#,##0.00'
    total_c.fill = PatternFill("solid", fgColor=_GREEN_LIGHT)
    total_c.border = _thin_border()
    total_c.alignment = Alignment(horizontal="center")

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    ws.freeze_panes = "A2"
